from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime
from typing import Dict

class RaporUretici:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_styles()
    
    def _setup_styles(self):
        """Rapor stilleri ayarla"""
        # Başlık stili
        self.styles['Title'].fontSize = 18
        self.styles['Title'].spaceAfter = 20
        
        # Alt başlık stili
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        ))
        
        # Metrik stili
        self.styles.add(ParagraphStyle(
            name='Metric',
            parent=self.styles['Normal'],
            fontSize=12,
            spaceAfter=8,
            leftIndent=20
        ))
    
    async def rapor_olustur(self, analiz_sonuclari: Dict, analiz_klasoru: str) -> Dict:
        """PDF ve Excel raporlarını oluştur"""
        
        # PDF raporu oluştur
        pdf_yolu = os.path.join(analiz_klasoru, "rapor.pdf")
        await self._pdf_rapor_olustur(analiz_sonuclari, pdf_yolu)
        
        # Excel raporu oluştur
        excel_yolu = os.path.join(analiz_klasoru, "rapor.xlsx")
        await self._excel_rapor_olustur(analiz_sonuclari, excel_yolu)
        
        return {
            'pdf_path': pdf_yolu,
            'excel_path': excel_yolu
        }
    
    async def _pdf_rapor_olustur(self, sonuclar: Dict, cikti_yolu: str):
        """PDF raporu oluştur"""
        doc = SimpleDocTemplate(cikti_yolu, pagesize=A4, topMargin=2*cm)
        story = []
        
        # Başlık
        story.append(Paragraph("Zeytin Ağacı Analiz Raporu", self.styles['Title']))
        story.append(Spacer(1, 12))
        
        # Tarih
        tarih = datetime.now().strftime('%d.%m.%Y %H:%M')
        story.append(Paragraph(f"Rapor Tarihi: {tarih}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Özet Bilgiler
        story.append(Paragraph("Özet Bilgiler", self.styles['Subtitle']))
        
        ozet_data = [
            ['Metrik', 'Değer'],
            ['Toplam Ağaç Sayısı', str(sonuclar.get('toplam_agac', 0))],
            ['Toplam Zeytin Sayısı', str(sonuclar.get('toplam_zeytin', 0))],
            ['Tahmini Zeytin Miktarı', f"{sonuclar.get('tahmini_zeytin_miktari', 0):.2f} kg"],
            ['Ortalama Ağaç Çapı', f"{sonuclar.get('agac_cap_ortalama', 0):.1f} cm"],
            ['NDVI Ortalama', f"{sonuclar.get('ndvi_ortalama', 0):.3f}"],
            ['GNDVI Ortalama', f"{sonuclar.get('gndvi_ortalama', 0):.3f}"],
            ['NDRE Ortalama', f"{sonuclar.get('ndre_ortalama', 0):.3f}"],
            ['Sağlık Durumu', sonuclar.get('saglik_durumu', 'Bilinmiyor')]
        ]
        
        ozet_table = Table(ozet_data, colWidths=[8*cm, 8*cm])
        ozet_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(ozet_table)
        story.append(Spacer(1, 20))
        
        # Detaylı Analiz
        if sonuclar.get('detaylar'):
            story.append(Paragraph("Detaylı Analiz Sonuçları", self.styles['Subtitle']))
            
            detay_data = [['Dosya', 'Ağaç Sayısı', 'Zeytin Sayısı', 'Ort. Çap (cm)']]
            
            for detay in sonuclar['detaylar']:
                detay_data.append([
                    detay['dosya'],
                    str(detay['agac_sayisi']),
                    str(detay['zeytin_sayisi']),
                    f"{detay['ortalama_cap']:.1f}"
                ])
            
            detay_table = Table(detay_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
            detay_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(detay_table)
            story.append(Spacer(1, 20))
        
        # Sağlık Değerlendirmesi
        story.append(Paragraph("Sağlık Değerlendirmesi", self.styles['Subtitle']))
        
        saglik_aciklama = self._saglik_aciklamasi_getir(sonuclar.get('saglik_durumu', ''))
        story.append(Paragraph(saglik_aciklama, self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Öneriler
        story.append(Paragraph("Öneriler", self.styles['Subtitle']))
        oneriler = self._oneri_getir(sonuclar)
        for oneri in oneriler:
            story.append(Paragraph(f"• {oneri}", self.styles['Normal']))
        
        # PDF'i oluştur
        doc.build(story)
    
    async def _excel_rapor_olustur(self, sonuclar: Dict, cikti_yolu: str):
        """Excel raporu oluştur"""
        wb = openpyxl.Workbook()
        
        # Özet sayfası
        ws_ozet = wb.active
        ws_ozet.title = "Özet"
        
        # Başlık
        ws_ozet['A1'] = "Zeytin Ağacı Analiz Raporu"
        ws_ozet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_ozet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_ozet.merge_cells('A1:B1')
        
        # Tarih
        ws_ozet['A3'] = "Rapor Tarihi:"
        ws_ozet['B3'] = datetime.now().strftime('%d.%m.%Y %H:%M')
        
        # Özet veriler
        ozet_veriler = [
            ("Toplam Ağaç Sayısı", sonuclar.get('toplam_agac', 0)),
            ("Toplam Zeytin Sayısı", sonuclar.get('toplam_zeytin', 0)),
            ("Tahmini Zeytin Miktarı (kg)", f"{sonuclar.get('tahmini_zeytin_miktari', 0):.2f}"),
            ("Ortalama Ağaç Çapı (cm)", f"{sonuclar.get('agac_cap_ortalama', 0):.1f}"),
            ("NDVI Ortalama", f"{sonuclar.get('ndvi_ortalama', 0):.3f}"),
            ("GNDVI Ortalama", f"{sonuclar.get('gndvi_ortalama', 0):.3f}"),
            ("NDRE Ortalama", f"{sonuclar.get('ndre_ortalama', 0):.3f}"),
            ("Sağlık Durumu", sonuclar.get('saglik_durumu', 'Bilinmiyor'))
        ]
        
        for i, (metrik, deger) in enumerate(ozet_veriler, start=5):
            ws_ozet[f'A{i}'] = metrik
            ws_ozet[f'B{i}'] = deger
            ws_ozet[f'A{i}'].font = Font(bold=True)
        
        # Detay sayfası
        if sonuclar.get('detaylar'):
            ws_detay = wb.create_sheet("Detaylar")
            
            # Başlıklar
            headers = ["Dosya", "Ağaç Sayısı", "Zeytin Sayısı", "Ortalama Çap (cm)"]
            for col, header in enumerate(headers, start=1):
                cell = ws_detay.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Veriler
            for row, detay in enumerate(sonuclar['detaylar'], start=2):
                ws_detay.cell(row=row, column=1, value=detay['dosya'])
                ws_detay.cell(row=row, column=2, value=detay['agac_sayisi'])
                ws_detay.cell(row=row, column=3, value=detay['zeytin_sayisi'])
                ws_detay.cell(row=row, column=4, value=round(detay['ortalama_cap'], 1))
            
            # Sütun genişliklerini ayarla
            for column in ws_detay.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_detay.column_dimensions[column_letter].width = adjusted_width
        
        # Excel dosyasını kaydet
        wb.save(cikti_yolu)
    
    def _saglik_aciklamasi_getir(self, saglik_durumu: str) -> str:
        """Sağlık durumu açıklaması"""
        aciklamalar = {
            "Çok Sağlıklı": "Ağaçlarınız çok iyi durumda. NDVI değerleri yüksek ve bitki sağlığı mükemmel.",
            "Sağlıklı": "Ağaçlarınız sağlıklı durumda. Normal bakım ve sulama ile devam edebilirsiniz.",
            "Orta Düzeyde Stresli": "Ağaçlarınızda hafif stres belirtileri var. Sulama ve gübreleme kontrolü önerilir.",
            "Stresli": "Ağaçlarınız stresli durumda. Acil sulama ve beslenme desteği gerekli.",
            "Çok Stresli/Hasta": "Ağaçlarınız ciddi sıkıntıda. Derhal uzman desteği alınması önerilir."
        }
        return aciklamalar.get(saglik_durumu, "Sağlık durumu değerlendirilemedi.")
    
    def _oneri_getir(self, sonuclar: Dict) -> list:
        """Analiz sonuçlarına göre öneriler"""
        oneriler = []
        
        ndvi = sonuclar.get('ndvi_ortalama', 0)
        agac_sayisi = sonuclar.get('toplam_agac', 0)
        
        if ndvi < 0.3:
            oneriler.append("Düşük NDVI değerleri nedeniyle sulama artırılması önerilir.")
            oneriler.append("Toprak analizi yaptırarak besленme eksikliği kontrol edilmelidir.")
        
        if ndvi > 0.7:
            oneriler.append("Yüksek NDVI değerleri sağlıklı bitki gelişimini göstermektedir.")
            oneriler.append("Mevcut bakım rutininizi sürdürün.")
        
        if agac_sayisi > 0:
            verimlilik = sonuclar.get('toplam_zeytin', 0) / agac_sayisi
            if verimlilik < 50:
                oneriler.append("Ağaç başına düşük zeytin sayısı tespit edildi. Budama ve gübreleme yapılması önerilir.")
        
        if not oneriler:
            oneriler.append("Genel olarak bahçeniz iyi durumda. Düzenli bakım ve kontrole devam edin.")
        
        return oneriler